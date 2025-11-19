"""
Модуль для экспорта данных в различные форматы (CSV, XLSX)
"""

import os
import pandas as pd
from typing import List, Dict
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Alignment
import shutil
import config


class DataExporter:
    """Экспортер данных в различные форматы"""
    
    def __init__(self, output_folder: str = None, template_path: str = None):
        self.output_folder = output_folder or config.OUTPUT_FOLDER
        self.template_path = template_path or self._find_template()
        
        # Создаем папку для вывода, если не существует
        os.makedirs(self.output_folder, exist_ok=True)
    
    def _find_template(self) -> str:
        """Ищет файл шаблона в директории проекта"""
        template_files = [
            'Таблица - шаблон обновленный.xlsx',
            'template.xlsx',
            'шаблон.xlsx'
        ]
        
        for template_file in template_files:
            if os.path.exists(template_file):
                return template_file
        
        return None
    
    def export(self, scenes: List[Dict], columns: List[str], 
               format: str = 'xlsx', filename: str = None) -> str:
        """
        Экспортирует сцены в указанный формат
        
        Args:
            scenes: список проанализированных сцен
            columns: список колонок для экспорта
            format: формат файла ('csv' или 'xlsx')
            filename: имя файла (опционально)
            
        Returns:
            str: путь к созданному файлу
        """
        # Создаем DataFrame
        df = self._create_dataframe(scenes, columns)
        
        # Генерируем имя файла, если не указано
        if not filename:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f'preproduction_{timestamp}'
        
        # Убираем расширение из имени файла, если есть
        filename_base = os.path.splitext(filename)[0]
        
        # Экспортируем в нужный формат
        if format.lower() == 'csv':
            output_path = os.path.join(self.output_folder, f'{filename_base}.csv')
            df.to_csv(output_path, index=False, encoding='utf-8-sig')
        elif format.lower() == 'xlsx':
            output_path = os.path.join(self.output_folder, f'{filename_base}.xlsx')
            self._export_to_excel(df, output_path)
        else:
            raise ValueError(f"Неподдерживаемый формат: {format}")
        
        return output_path
    
    def _create_dataframe(self, scenes: List[Dict], columns: List[str]) -> pd.DataFrame:
        """
        Создает DataFrame из сцен с указанными колонками
        
        Args:
            scenes: список сцен
            columns: список названий колонок
            
        Returns:
            pd.DataFrame: готовый DataFrame
        """
        data = []
        
        # Маппинг названий колонок на поля в данных
        column_mapping = {
            'Номер сцены': 'scene_number',
            'Локация': 'location',
            'Внутри/Снаружи': 'interior_exterior',
            'Время суток': 'time_of_day',
            'Персонажи': 'characters',
            'Персонажи (основные)': 'main_characters',
            'Персонажи (второстепенные)': 'secondary_characters',
            'Массовка': 'extras',
            'Реквизит': 'props',
            'Транспорт': 'transport',
            'Животные': 'animals',
            'Спецэффекты': 'sfx',
            'Грим/Костюмы': 'makeup',
            'Каскадеры': 'stunts',
            'Музыка/Звук': 'music',
            'Оружие': 'weapons',
            'Еда/Напитки': 'food',
            'Краткое описание': 'description',
            'Количество слов': 'word_count'
        }
        
        for scene in scenes:
            row = {}
            
            for column in columns:
                field = column_mapping.get(column, column.lower().replace(' ', '_'))
                
                value = scene.get(field, '')
                
                # Форматируем значение
                if isinstance(value, list):
                    if value:
                        value = ', '.join(str(v) for v in value)
                    else:
                        value = ''
                elif value is None:
                    value = ''
                
                row[column] = value
            
            data.append(row)
        
        return pd.DataFrame(data)
    
    def _export_to_excel(self, df: pd.DataFrame, output_path: str):
        """
        Экспортирует DataFrame в Excel с форматированием используя шаблон
        
        Args:
            df: DataFrame для экспорта
            output_path: путь к выходному файлу
        """
        # Если шаблон найден, используем его
        if self.template_path and os.path.exists(self.template_path):
            self._export_with_template(df, output_path)
        else:
            # Используем стандартное форматирование
            self._export_with_default_style(df, output_path)
    
    def _export_with_template(self, df: pd.DataFrame, output_path: str):
        """Экспорт с использованием шаблона"""
        try:
            # Загружаем шаблон
            wb = load_workbook(self.template_path)
            ws = wb.active
            
            # Маппинг колонок DataFrame на колонки шаблона
            template_columns = {
                'Номер сцены': 'Сцена',
                'Локация': 'Объект',
                'Внутри/Снаружи': 'Инт / нат',
                'Время суток': 'Режим',
                'Персонажи': 'Персонажи',
                'Персонажи (основные)': 'Персонажи',
                'Персонажи (второстепенные)': 'Персонажи',
                'Массовка': 'Массовка',
                'Реквизит': 'Реквизит',
                'Транспорт': 'Игровой транспорт',
                'Животные': 'Животные',
                'Спецэффекты': 'Трюк',
                'Грим/Костюмы': 'Грим',
                'Каскадеры': 'Трюк',
                'Костюмы': 'Костюм',
                'Краткое описание': 'Синопсис'
            }
            
            # Читаем заголовки из шаблона (строка 2)
            template_headers = []
            for cell in ws[2]:
                template_headers.append(cell.value)
            
            print(f"Заголовки шаблона: {template_headers}")
            print(f"Колонки DataFrame: {list(df.columns)}")
            
            # Создаем маппинг индексов колонок (нумерация с 1)
            column_map = {}
            for df_col in df.columns:
                template_col = template_columns.get(df_col, None)
                if template_col and template_col in template_headers:
                    # Находим позицию колонки (с учетом что индексы начинаются с 1)
                    col_index = template_headers.index(template_col) + 1
                    column_map[df_col] = col_index
                    print(f"Маппинг: '{df_col}' → '{template_col}' (колонка {col_index})")
            
            if not column_map:
                print("Внимание: не найдено соответствий между данными и шаблоном")
                raise ValueError("Не удалось сопоставить колонки с шаблоном")
            
            # Удаляем пустые строки из шаблона (оставляем только заголовки)
            # Удаляем строки с 3 по max_row
            if ws.max_row > 2:
                ws.delete_rows(3, ws.max_row - 2)
            
            # Начинаем записывать данные с 3-й строки (после заголовков)
            start_row = 3
            
            # Получаем образец стиля из строки заголовков для копирования границ
            header_row = 2
            
            print(f"\nЗаписываем {len(df)} сцен в шаблон...")
            
            for idx, (_, row_data) in enumerate(df.iterrows()):
                row_num = start_row + idx
                
                # Копируем стили границ из заголовков для всех колонок
                for col_idx in range(1, ws.max_column + 1):
                    header_cell = ws.cell(row=header_row, column=col_idx)
                    target_cell = ws.cell(row=row_num, column=col_idx)
                    
                    # Копируем только границы и базовое выравнивание
                    if header_cell.border:
                        target_cell.border = header_cell.border.copy()
                    
                    # Устанавливаем базовое выравнивание
                    target_cell.alignment = Alignment(
                        horizontal='left',
                        vertical='center',
                        wrap_text=True
                    )
                
                # Записываем данные в соответствующие колонки
                for df_col, template_col_idx in column_map.items():
                    cell = ws.cell(row=row_num, column=template_col_idx)
                    value = row_data[df_col]
                    
                    # Форматируем значение
                    if pd.isna(value) or value == '':
                        cell.value = ''
                    else:
                        cell.value = str(value)
                    
                    print(f"  Строка {row_num}, Колонка {template_col_idx} ({df_col}): {str(value)[:50]}")
            
            # Сохраняем
            wb.save(output_path)
            print(f"✓ Файл сохранен с использованием шаблона: {output_path}")
            
        except Exception as e:
            print(f"Ошибка при использовании шаблона: {e}")
            print("Используем стандартное форматирование...")
            self._export_with_default_style(df, output_path)
    
    def _export_with_default_style(self, df: pd.DataFrame, output_path: str):
        """Экспорт со стандартным форматированием"""
        # Создаем Excel writer
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Препродакшн', index=False)
            
            # Получаем workbook и worksheet для форматирования
            workbook = writer.book
            worksheet = writer.sheets['Препродакшн']
            
            # Форматируем заголовки
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            
            header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            header_font = Font(color='FFFFFF', bold=True, size=11)
            header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            # Стиль границ
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Применяем стиль к заголовкам
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
                cell.border = thin_border
            
            # Форматируем ячейки данных
            cell_alignment = Alignment(vertical='top', wrap_text=True)
            
            for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, 
                                          min_col=1, max_col=worksheet.max_column):
                for cell in row:
                    cell.alignment = cell_alignment
                    cell.border = thin_border
            
            # Автоматически настраиваем ширину колонок
            self._auto_adjust_column_width(worksheet, df)
            
            # Закрепляем первую строку
            worksheet.freeze_panes = 'A2'
    
    def _auto_adjust_column_width(self, worksheet, df: pd.DataFrame):
        """
        Автоматически настраивает ширину колонок
        
        Args:
            worksheet: worksheet объект openpyxl
            df: DataFrame с данными
        """
        for idx, column in enumerate(df.columns, 1):
            # Находим максимальную длину в колонке
            max_length = len(str(column))
            
            for value in df[column]:
                if value:
                    # Учитываем многострочные значения
                    lines = str(value).split('\n')
                    max_line_length = max(len(line) for line in lines)
                    max_length = max(max_length, min(max_line_length, 50))
            
            # Устанавливаем ширину с небольшим запасом
            adjusted_width = min(max_length + 2, 60)
            column_letter = worksheet.cell(row=1, column=idx).column_letter
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    def export_summary(self, scenes: List[Dict], filename: str = None) -> str:
        """
        Экспортирует сводную статистику по сценарию
        
        Args:
            scenes: список проанализированных сцен
            filename: имя файла (опционально)
            
        Returns:
            str: путь к созданному файлу
        """
        # Собираем статистику
        summary = self._collect_summary_stats(scenes)
        
        # Генерируем имя файла
        if not filename:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f'summary_{timestamp}.xlsx'
        
        output_path = os.path.join(self.output_folder, filename)
        
        # Создаем Excel с несколькими листами
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            # Лист 1: Общая статистика
            general_df = pd.DataFrame([
                ['Всего сцен', summary['total_scenes']],
                ['Всего персонажей', summary['total_characters']],
                ['Всего локаций', summary['total_locations']],
                ['Средняя длина сцены (слов)', summary['avg_scene_length']],
            ], columns=['Показатель', 'Значение'])
            general_df.to_excel(writer, sheet_name='Общая статистика', index=False)
            
            # Лист 2: Персонажи
            if summary['characters']:
                characters_df = pd.DataFrame(
                    list(summary['characters'].items()),
                    columns=['Персонаж', 'Количество сцен']
                ).sort_values('Количество сцен', ascending=False)
                characters_df.to_excel(writer, sheet_name='Персонажи', index=False)
            
            # Лист 3: Локации
            if summary['locations']:
                locations_df = pd.DataFrame(
                    list(summary['locations'].items()),
                    columns=['Локация', 'Количество сцен']
                ).sort_values('Количество сцен', ascending=False)
                locations_df.to_excel(writer, sheet_name='Локации', index=False)
            
            # Лист 4: Распределение по времени суток
            if summary['time_of_day']:
                time_df = pd.DataFrame(
                    list(summary['time_of_day'].items()),
                    columns=['Время суток', 'Количество сцен']
                )
                time_df.to_excel(writer, sheet_name='Время суток', index=False)
        
        return output_path
    
    def _collect_summary_stats(self, scenes: List[Dict]) -> Dict:
        """Собирает сводную статистику"""
        from collections import Counter
        
        total_scenes = len(scenes)
        
        # Подсчитываем персонажей
        all_characters = []
        for scene in scenes:
            all_characters.extend(scene.get('characters', []))
        character_counts = Counter(all_characters)
        
        # Подсчитываем локации
        all_locations = [scene.get('location', '') for scene in scenes if scene.get('location')]
        location_counts = Counter(all_locations)
        
        # Подсчитываем время суток
        all_times = [scene.get('time_of_day', '') for scene in scenes if scene.get('time_of_day')]
        time_counts = Counter(all_times)
        
        # Средняя длина сцены
        word_counts = [scene.get('word_count', 0) for scene in scenes]
        avg_length = sum(word_counts) / len(word_counts) if word_counts else 0
        
        return {
            'total_scenes': total_scenes,
            'total_characters': len(character_counts),
            'total_locations': len(location_counts),
            'avg_scene_length': round(avg_length, 1),
            'characters': dict(character_counts),
            'locations': dict(location_counts),
            'time_of_day': dict(time_counts)
        }

